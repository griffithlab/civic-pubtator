#!/usr/bin/env python3
import argparse, os, shutil, subprocess, sys, tempfile

# ── LibreOffice detection ─────────────────────────────────────────────────────

def find_soffice():
    path = shutil.which("soffice")
    if path:
        return path
    mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.isfile(mac_path):
        return mac_path
    return None

def soffice_convert(soffice, src, out_dir):
    """Convert src to PDF in out_dir via soffice. Returns path of created PDF."""
    result = subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", out_dir, src],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip())
    stem = os.path.splitext(os.path.basename(src))[0]
    return os.path.join(out_dir, stem + ".pdf")

# ── Per-type processors ───────────────────────────────────────────────────────

def process_pdf(src, stem, s_dir):
    out_dir = os.path.join(s_dir, stem)
    os.makedirs(out_dir, exist_ok=True)
    dst = os.path.join(out_dir, stem + ".pdf")
    shutil.copy2(src, dst)
    print(f"  Copied → {dst}")

def process_word(src, stem, s_dir, soffice):
    out_dir = os.path.join(s_dir, stem)
    os.makedirs(out_dir, exist_ok=True)

    if soffice:
        created = soffice_convert(soffice, src, out_dir)
        # soffice names output after the source stem; rename to match our convention
        dst = os.path.join(out_dir, stem + ".pdf")
        if created != dst:
            os.replace(created, dst)
        print(f"  Word → PDF (soffice): {dst}")
        return

    # ── Fallback: python-docx + reportlab (.docx only) ───────────────────────
    ext = os.path.splitext(src)[1].lower()
    if ext == ".doc":
        print("  WARNING: .doc requires LibreOffice for conversion — skipping.")
        print("           Install with: brew install --cask libreoffice")
        return
    try:
        from docx import Document
    except ImportError:
        sys.exit("ERROR: python-docx not installed. Run: pip3 install python-docx")
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        sys.exit("ERROR: reportlab not installed. Run: pip3 install reportlab")

    dst = os.path.join(out_dir, stem + ".pdf")
    doc = Document(src)
    styles = getSampleStyleSheet()
    story = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if text:
            story.append(Paragraph(text, styles["Normal"]))
            story.append(Spacer(1, 4))
    SimpleDocTemplate(dst, pagesize=letter).build(story)
    print(f"  Word → PDF (reportlab fallback): {dst}")

def process_excel(src, stem, s_dir, soffice):
    try:
        import openpyxl
    except ImportError:
        sys.exit("ERROR: openpyxl not installed. Run: pip3 install openpyxl")

    MAX_ROWS = 1000
    wb = openpyxl.load_workbook(src, data_only=True)

    for tab_num, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_ROWS:
                break
            rows.append(row)

        if not rows:
            print(f"  Sheet '{sheet_name}' (tab {tab_num}): empty, skipping")
            continue

        tab_label = f"tab_{tab_num:02d}"
        out_dir = os.path.join(s_dir, stem, tab_label)
        os.makedirs(out_dir, exist_ok=True)
        dst = os.path.join(out_dir, f"{stem}.pdf")

        if soffice:
            from openpyxl.utils import get_column_letter
            tmp_dir = tempfile.mkdtemp()
            try:
                tmp_wb = openpyxl.Workbook()
                tmp_ws = tmp_wb.active
                tmp_ws.title = sheet_name
                for row in rows:
                    tmp_ws.append([v if v is not None else "" for v in row])

                # Auto-size columns based on content
                ncols = max(len(r) for r in rows)
                for col_idx in range(1, ncols + 1):
                    max_len = max(
                        len(str(r[col_idx - 1])) if col_idx <= len(r) and r[col_idx - 1] is not None else 0
                        for r in rows
                    )
                    tmp_ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(max_len + 2, 50))

                # Landscape, A3, fit all columns onto one page wide
                tmp_ws.page_setup.orientation = "landscape"
                tmp_ws.page_setup.paperSize = 8  # A3
                tmp_ws.page_setup.fitToWidth = 1
                tmp_ws.page_setup.fitToHeight = 0
                tmp_ws.sheet_properties.pageSetUpPr.fitToPage = True

                # Header: sheet name centred, tab number on the right
                tmp_ws.oddHeader.center.text = sheet_name
                tmp_ws.oddHeader.right.text  = f"Tab {tab_num}"

                tmp_xlsx = os.path.join(tmp_dir, f"{stem}.xlsx")
                tmp_wb.save(tmp_xlsx)
                created = soffice_convert(soffice, tmp_xlsx, tmp_dir)
                shutil.move(created, dst)
            finally:
                shutil.rmtree(tmp_dir, ignore_errors=True)
            print(f"  Sheet '{sheet_name}' (tab {tab_num}, {len(rows)} rows) → {dst} (soffice)")

        else:
            # ── Fallback: reportlab ───────────────────────────────────────────
            try:
                from reportlab.lib import colors
                from reportlab.lib.pagesizes import landscape, letter
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib.units import inch
                from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
            except ImportError:
                sys.exit("ERROR: reportlab not installed. Run: pip3 install reportlab")

            MARGINS = 0.5 * inch
            MIN_COL_WIDTH = 1.3 * inch

            str_rows = [[str(v) if v is not None else "" for v in row] for row in rows]
            ncols = max(len(r) for r in str_rows)
            str_rows = [r + [""] * (ncols - len(r)) for r in str_rows]

            col_width  = max(MIN_COL_WIDTH, landscape(letter)[0] / ncols)
            page_width = max(landscape(letter)[0], ncols * col_width + 2 * MARGINS)
            page_height = landscape(letter)[1]
            font_size  = max(5, min(8, int(col_width / 0.12)))

            styles = getSampleStyleSheet()
            title = Paragraph(f"<b>{sheet_name}</b>  <font size='8'>(Tab {tab_num})</font>",
                              styles["Normal"])

            table = Table(str_rows, colWidths=[col_width] * ncols, repeatRows=1)
            table.setStyle(TableStyle([
                ("BACKGROUND",     (0, 0), (-1,  0), colors.grey),
                ("TEXTCOLOR",      (0, 0), (-1,  0), colors.whitesmoke),
                ("FONTSIZE",       (0, 0), (-1, -1), font_size),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ("GRID",           (0, 0), (-1, -1), 0.25, colors.black),
                ("VALIGN",         (0, 0), (-1, -1), "TOP"),
                ("WORDWRAP",       (0, 0), (-1, -1), True),
            ]))
            SimpleDocTemplate(dst, pagesize=(page_width, page_height),
                              leftMargin=MARGINS, rightMargin=MARGINS,
                              topMargin=MARGINS, bottomMargin=MARGINS).build(
                [title, Spacer(1, 0.15 * inch), table])
            print(f"  Sheet '{sheet_name}' (tab {tab_num}, {len(rows)} rows) → {dst} (reportlab fallback)")

    wb.close()

# ── Main ──────────────────────────────────────────────────────────────────────

IGNORED_FILES = {".DS_Store"}

def main():
    parser = argparse.ArgumentParser(
        description="Convert supplementary files in <input_dir>/s/ to PDFs."
    )
    parser.add_argument("input_dir",
                        help="Source directory containing an s/ subdirectory")
    parser.add_argument("--no-libreoffice", action="store_true",
                        help="Use the reportlab/python-docx fallback even if LibreOffice is installed")
    args = parser.parse_args()

    if args.no_libreoffice:
        soffice = None
        print("LibreOffice disabled by --no-libreoffice; using reportlab/python-docx fallback.")
    else:
        soffice = find_soffice()
        if soffice:
            print(f"LibreOffice found: {soffice}")
        else:
            print("WARNING: LibreOffice (soffice) not found.")
            print("         Word and Excel conversion will use a basic fallback with reduced fidelity.")
            print("         For best results install LibreOffice: brew install --cask libreoffice\n")

    input_dir = os.path.abspath(args.input_dir)
    s_dir = os.path.join(input_dir, "s")

    if not os.path.isdir(s_dir):
        print(f"No s/ subdirectory found in {input_dir}, nothing to do.")
        sys.exit(0)

    for fname in sorted(os.listdir(s_dir)):
        fpath = os.path.join(s_dir, fname)
        if not os.path.isfile(fpath) or fname in IGNORED_FILES or fname.startswith("~$"):
            continue

        stem = os.path.splitext(fname)[0]
        ext  = os.path.splitext(fname)[1].lower()

        print(f"\nProcessing: {fname}")

        if ext == ".pdf":
            process_pdf(fpath, stem, s_dir)
        elif ext in (".docx", ".doc"):
            process_word(fpath, stem, s_dir, soffice)
        elif ext in (".xlsx", ".xls"):
            process_excel(fpath, stem, s_dir, soffice)
        else:
            print(f"  Unsupported type ({ext}), skipping.")

if __name__ == "__main__":
    main()
